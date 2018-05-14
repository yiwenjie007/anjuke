# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import Workbook
import pymysql

class AnjukePipeline(object):

    def __init__(self):
        self.excel_file = 'house.xlsx'
        self.item_list = []
        self.header=['楼盘名称', '楼盘售价(元/㎡)', '周边价(元/㎡)', '户型', '楼盘地址', '电话', '开盘时间', '交房时间', '链接地址']

    def open_spider(self, spider):
        self.wb = Workbook()
        self.sheet = self.wb.create_sheet('楼盘信息', 0)
        for i in range(1, len(self.header) + 1):
            self.sheet.cell(row=1, column=i, value=self.header[i-1])

    def close_spider(self, spider):
        for i in range(2, len(self.item_list) + 2):
            for j in range(1, len(self.header) + 1):
                self.sheet.cell(row=i, column=j, value=self.get_value(dict(self.item_list[i-2]), j))
        self.wb.save(self.excel_file)
        self.wb.close()

    def process_item(self, item, spider):
        self.item_list.append(item)
        return item

    def get_value(self, item, j):
        if j == 1:
            return None if not 'title' in item.keys() else item['title']
        if j == 2:
            return None if not 'price' in item.keys() else item['price']
        if j == 3:
            return None if not 'around_price' in item.keys() else item['around_price']
        if j == 4:
            return None if not 'house_type' in item.keys() else item['house_type']
        if j == 5:
            return None if not 'address' in item.keys() else item['address']
        if j == 6:
            return None if not 'phone' in item.keys() else item['phone']
        if j == 7:
            return None if not 'opentime' in item.keys() else item['opentime']
        if j == 8:
            return None if not 'completetime' in item.keys() else item['completetime']
        if j == 9:
            return None if not 'url' in item.keys() else item['url']
			
			
class AnjuKeMySqlPipeline(object):
    def __init__(self, host, port, user, password, database):
        self.host = host
        self.port = port
        self.user = user
        self.password = password
        self.database = database

    @classmethod
    def from_crawler(cls, crawler):
        return cls(
            host=crawler.settings.get('MYSQL_HOST'),
            port=crawler.settings.get('MYSQL_PORT'),
            user=crawler.settings.get('MYSQL_USER'),
            password=crawler.settings.get('MYSQL_PASSWORD'),
            database=crawler.settings.get('MYSQL_DB')
        )

    def open_spider(self, spider):
        self.db = pymysql.connect(host=self.host, port=self.port, user=self.user, password=self.password, database=self.database, charset='utf8')
        self.cursor = self.db.cursor()

    def close_spider(self, spider):
        self.cursor.close()
        self.db.close()

    def process_item(self, item, spider):
        data = dict(item)
        keys = ', '.join(data.keys())
        values = ', '.join(['%s'] * len(data))
        sql = 'insert into {table}({keys}) values ({values})'.format(table=item.collection, keys=keys, values=values)
        try:
            self.cursor.execute(sql, tuple(data.values()))
            self.db.commit()
        except:
            self.db.rollback()
        return item
